import io
import json as _json
import os

import dynamo
import output_naming
import s3
from logger import get_logger, log_event

log = get_logger(__name__)


def _normalize_cell(value):
    return "" if value is None else str(value)


def _rows_from_words(words: list[dict], y_tolerance: float = 5.0) -> list[list[str]]:
    """Group words into rows by Y position. Simple sequential output."""
    rows: list[list[str]] = []
    current_bucket: int | None = None
    current_row: list[tuple[float, str]] = []

    for word in sorted(words, key=lambda item: (float(item.get("top", 0.0)), float(item.get("x0", 0.0)))):
        bucket = int(round(float(word.get("top", 0.0)) / y_tolerance))
        if current_bucket is None:
            current_bucket = bucket
        if bucket != current_bucket and current_row:
            rows.append([text for _x, text in sorted(current_row, key=lambda item: item[0])])
            current_row = []
            current_bucket = bucket
        current_row.append((float(word.get("x0", 0.0)), _normalize_cell(word.get("text"))))

    if current_row:
        rows.append([text for _x, text in sorted(current_row, key=lambda item: item[0])])

    return rows


def _cluster_columns(x_positions: list[float], x_tolerance: float = 15.0) -> list[float]:
    """Cluster X positions into column boundaries. Returns sorted cluster centres."""
    if not x_positions:
        return []
    sorted_xs = sorted(x_positions)
    clusters: list[list[float]] = [[sorted_xs[0]]]
    for x in sorted_xs[1:]:
        if x - clusters[-1][-1] <= x_tolerance:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    return [sum(c) / len(c) for c in clusters]


def _assign_column(x0: float, col_centres: list[float]) -> int:
    """Return the index of the nearest column centre."""
    best_idx = 0
    best_dist = abs(x0 - col_centres[0])
    for i, cx in enumerate(col_centres[1:], start=1):
        d = abs(x0 - cx)
        if d < best_dist:
            best_dist = d
            best_idx = i
    return best_idx


def _rows_from_words_columnar(words: list[dict], y_tolerance: float = 5.0,
                               x_tolerance: float = 15.0) -> list[list[str]]:
    """Group words into rows with proper column alignment using X clustering."""
    if not words:
        return []

    # Build column centres from all x0 values
    all_x0 = [float(w.get("x0", 0.0)) for w in words]
    col_centres = _cluster_columns(all_x0, x_tolerance)
    n_cols = len(col_centres)
    if n_cols == 0:
        return _rows_from_words(words, y_tolerance)

    # Group words into rows by Y bucket
    sorted_words = sorted(words, key=lambda w: (float(w.get("top", 0.0)), float(w.get("x0", 0.0))))
    row_buckets: list[list[dict]] = []
    current_bucket: int | None = None
    current_group: list[dict] = []

    for w in sorted_words:
        bucket = int(round(float(w.get("top", 0.0)) / y_tolerance))
        if current_bucket is None:
            current_bucket = bucket
        if bucket != current_bucket and current_group:
            row_buckets.append(current_group)
            current_group = []
            current_bucket = bucket
        current_group.append(w)
    if current_group:
        row_buckets.append(current_group)

    # Build output rows with column assignment
    rows: list[list[str]] = []
    for group in row_buckets:
        row = [""] * n_cols
        for w in group:
            col_idx = _assign_column(float(w.get("x0", 0.0)), col_centres)
            existing = row[col_idx]
            text = _normalize_cell(w.get("text"))
            row[col_idx] = f"{existing} {text}".strip() if existing else text
        rows.append(row)

    return rows


def _detect_table_start(rows: list[list[str]], n_cols: int) -> int:
    """Find the first row where the clean table begins.

    Strategy: scan rows top-down. A "clean" table row uses most of the columns.
    The header region of irregular documents (like Brazilian payslips) has rows
    where words are scattered across fewer columns than the main table.

    Returns the index of the first row that starts the consistent table region.
    """
    if not rows or n_cols <= 1:
        return 0

    # A row is "full" if it populates >= 60% of columns
    threshold = max(2, int(n_cols * 0.6))

    # Find the first stretch of 2+ consecutive "full" rows — that's the table
    consecutive = 0
    start_candidate = 0
    for i, row in enumerate(rows):
        filled = sum(1 for cell in row if cell.strip())
        if filled >= threshold:
            if consecutive == 0:
                start_candidate = i
            consecutive += 1
            if consecutive >= 2:
                return start_candidate
        else:
            consecutive = 0

    return 0


def _page_rows(page) -> list[list[str]]:
    tables = page.extract_tables() or []
    if tables:
        rows: list[list[str]] = []
        for index, table in enumerate(tables):
            if index > 0:
                rows.append([])
            for row in table or []:
                rows.append([_normalize_cell(cell) for cell in row or []])
        return rows

    words = page.extract_words(keep_blank_chars=True, use_text_flow=True) or []
    return _rows_from_words(words)


def _non_empty_row_count(rows: list[list[str]]) -> int:
    return sum(1 for row in rows if any(str(cell or "").strip() for cell in row))


def _page_is_scanned(pdf_bytes: bytes, page_index_0based: int) -> bool:
    """Detect if a page is a scanned image (no extractable text, has images)."""
    import pymupdf

    doc = pymupdf.open(stream=pdf_bytes, filetype="pdf")
    try:
        page = doc.load_page(page_index_0based)
        text = (page.get_text("text") or "").strip()
        images = page.get_images(full=False)
        has_text = len(text) > 10
        has_images = len(images) > 0
        return not has_text and has_images
    finally:
        doc.close()


_OCR_MAX_PAGES = int(os.environ.get("OCR_MAX_PAGES_PER_JOB", "30"))


def _build_workbook(pdf_bytes: bytes, *, ocr_page_indices: list[int] | None = None, job: dict | None = None, job_id: str = "") -> tuple[bytes, list[str]]:
    import pdfplumber
    from openpyxl import Workbook

    workbook = Workbook()
    consolidated = workbook.active
    consolidated.title = "Consolidated"
    consolidated_row = 1
    last_header: list[str] | None = None
    warnings: list[str] = []
    ocr_set = set(ocr_page_indices or [])
    ocr_pages_used = 0

    log_event("info", "build_workbook_start", job,
              ocr_page_indices=list(ocr_set), ocr_set_size=len(ocr_set))

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        pages = list(pdf.pages)
        log_event("info", "pdfplumber_opened", job, page_count=len(pages))

        total_pages = len(pages)
        if not pages:
            workbook.create_sheet("Page 1")
        for page_index, page in enumerate(pages, start=1):
            page_0 = page_index - 1
            if job_id and total_pages > 0:
                pct = int(page_index / total_pages * 90)  # reserve 10% for save
                dynamo.update_progress(job_id, pct, f"Page {page_index}/{total_pages}")
            sheet = workbook.create_sheet()
            sheet.title = f"Page {page_index}"
            try:
                rows = _page_rows(page)
            except Exception as exc:
                warnings.append(f"Page {page_index} could not be extracted: {exc}")
                log_event("warning", "page_extract_error", job,
                          page_index=page_0, error=str(exc))
                rows = []

            non_empty = _non_empty_row_count(rows)
            log_event("info", "page_extract_result", job,
                      page_index=page_0, total_rows=len(rows),
                      non_empty_rows=non_empty, in_ocr_set=page_0 in ocr_set)

            # Self-sufficient OCR fallback: always try OCR for empty pages.
            # Upstream ocr_page_indices is a priority hint, not a gate.
            if non_empty == 0 and ocr_pages_used < _OCR_MAX_PAGES:
                trigger = None
                if page_0 in ocr_set:
                    trigger = "upstream"
                elif _page_is_scanned(pdf_bytes, page_0):
                    trigger = "inline_detection"

                if trigger:
                    ocr_pages_used += 1
                    log_event("info", "ocr_started", job,
                              page_index=page_0, trigger=trigger,
                              ocr_pages_used=ocr_pages_used)
                    try:
                        from pdf_ocr_pages import ocr_pdf_pages
                        ocr_results = ocr_pdf_pages(pdf_bytes, page_indices=[page_0])
                        r = ocr_results[0] if ocr_results else None
                        log_event("info", "ocr_result", job,
                                  page_index=page_0,
                                  source=r.source if r else "none",
                                  word_count=len(r.words) if r else 0,
                                  line_count=len(r.lines) if r else 0,
                                  first_words=[w.text for w in (r.words if r else [])[:5]])
                        if r and r.words:
                            ocr_words = [w.to_pdfplumber_dict() for w in r.words]
                            rows = _rows_from_words_columnar(ocr_words)
                            # Skip irregular header region
                            n_cols = max((len(row) for row in rows), default=0)
                            table_start = _detect_table_start(rows, n_cols)
                            if table_start > 0:
                                log_event("info", "header_skipped", job,
                                          page_index=page_0,
                                          skipped_rows=table_start,
                                          total_rows=len(rows))
                                rows = rows[table_start:]
                            non_empty = _non_empty_row_count(rows)
                            sheet.title = f"Page {page_index} (OCR)"
                            warnings.append(
                                f"Page {page_index} extracted via OCR "
                                f"({r.source}, {len(r.words)} words)."
                            )
                            log_event("info", "ocr_succeeded", job,
                                      page_index=page_0, rows=len(rows),
                                      non_empty_rows=non_empty)
                        else:
                            log_event("warning", "ocr_no_words", job,
                                      page_index=page_0,
                                      source=r.source if r else "none")
                            warnings.append(f"Page {page_index} OCR returned no words.")
                    except Exception as ocr_exc:
                        log_event("error", "ocr_failed", job,
                                  page_index=page_0, error=str(ocr_exc))
                        warnings.append(f"Page {page_index} OCR failed: {ocr_exc}")

            for row_index, row in enumerate(rows, start=1):
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index, column=col_index, value=value)

            if not rows:
                continue
            first_row = rows[0] if rows else []
            has_text_header = any(str(cell or "").strip() for cell in first_row)
            if has_text_header:
                last_header = [str(cell or "") for cell in first_row]
            elif last_header:
                rows = [last_header] + rows
            elif _non_empty_row_count(rows) <= 1:
                warnings.append(f"Page {page_index} has insufficient data for consolidated rows.")
                continue

            for row in rows:
                for col_index, value in enumerate(row, start=1):
                    consolidated.cell(row=consolidated_row, column=col_index, value=value)
                consolidated_row += 1

    log_event("info", "build_workbook_done", job,
              consolidated_rows=consolidated_row - 1,
              ocr_pages_used=ocr_pages_used,
              warning_count=len(warnings))
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue(), warnings


def _output_filename(body: dict, file_key: str) -> str:
    return output_naming.output_filename("pdf_to_xls", body, file_key, "xlsx")


def handler(event, context):
    body = _json.loads(event["Records"][0]["body"])
    job_id = body["job_id"]
    file_key = body["file_key"]
    analysis_result = body.get("analysis_result") or {}
    ocr_page_indices = analysis_result.get("ocr_page_indices") or []

    job = {
        "job_id": job_id,
        "operation": "pdf_to_xls",
        "file_key": file_key,
        "file_size_bytes": body.get("file_size_bytes", 0),
        "file_name": body.get("file_name", ""),
        "session_id": body.get("session_id", ""),
        "user_id": body.get("user_id", ""),
    }

    log_event("info", "job_started", job,
              has_analysis_result=bool(analysis_result),
              needs_ocr=analysis_result.get("needs_ocr"),
              ocr_page_indices=ocr_page_indices,
              recommendation=analysis_result.get("recommendation"),
              body_keys=list(body.keys()))

    try:
        dynamo.update_job(job_id, status="PROCESSING")
        data = s3.get_bytes(file_key)
        log_event("info", "file_loaded", job, size_bytes=len(data))

        dynamo.update_progress(job_id, 5, "Reading PDF")
        result, warnings = _build_workbook(data, ocr_page_indices=ocr_page_indices, job=job, job_id=job_id)
        out_key = s3.make_output_key(job_id, file_key, _output_filename(body, file_key))
        s3.put_bytes(out_key, result)

        log_event("info", "job_completed", job,
                  output_size=len(result), warning_count=len(warnings),
                  warnings=warnings)
        if warnings:
            dynamo.update_job(job_id, job_warnings=warnings)
        dynamo.mark_done(job_id, out_key)
    except Exception as exc:
        log_event("error", "job_failed", job, error=str(exc))
        log.exception("pdf_to_xls failed: %s", exc)
        dynamo.mark_failed(job_id, str(exc))
        raise
