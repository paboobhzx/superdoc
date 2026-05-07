"""Fast xlsx-to-PDF handler using openpyxl + reportlab.

Trades fidelity for speed: renders sheet rows as a grid table PDF.
Conditional formatting, charts, merged cells, and formulas are not preserved
— use the high-fidelity (LibreOffice) path for those.
"""
import io
import json as _json

import dynamo
import page_sizes
import s3
from logger import get_logger

log = get_logger(__name__)


def _xlsx_to_pdf(
    xlsx_bytes: bytes,
    paper_size: str | None = None,
    sheet_name: str | None = None,
) -> bytes:
    from openpyxl import load_workbook
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

    pw, ph = page_sizes.get(paper_size)

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0]
        rows = [
            [str(cell) if cell is not None else "" for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=(pw, ph), leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    story = []

    if rows:
        t = Table(rows, repeatRows=1)
        t.setStyle(TableStyle([
            ("GRID",           (0, 0), (-1, -1), 0.4, colors.grey),
            ("FONTSIZE",       (0, 0), (-1, -1), 8),
            ("BACKGROUND",     (0, 0), (-1, 0),  colors.lightgrey),
            ("FONTNAME",       (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.Color(0.97, 0.97, 0.97)]),
            ("VALIGN",         (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING",    (0, 0), (-1, -1), 4),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 4),
        ]))
        story.append(t)

    doc.build(story)
    return buf.getvalue()


def handler(event, context):
    if event.get("_warmup"):
        log.info("warmup ping received")
        return
    body = _json.loads(event["Records"][0]["body"])
    job_id = body["job_id"]
    file_key = body["file_key"]
    params = body.get("params") or {}
    paper_size = params.get("paper_size") or None
    sheet_name = params.get("sheet") or None

    try:
        dynamo.update_job(job_id, status="PROCESSING")
        data = s3.get_bytes(file_key)
        result = _xlsx_to_pdf(data, paper_size=paper_size, sheet_name=sheet_name)
        out_key = s3.make_output_key(job_id, file_key, "output.pdf")
        s3.put_bytes(out_key, result)
        dynamo.mark_done(job_id, out_key)
        log.info("xlsx_to_pdf_fast done", extra={"job_id": job_id})
    except Exception as exc:
        log.exception("xlsx_to_pdf_fast failed: %s", exc)
        dynamo.mark_failed(job_id, str(exc))
        raise
