# handlers/xlsx_to_csv.py - Converts a sheet of an .xlsx into text-like
# formats that are cheap to render without LibreOffice.
#
# The existing Lambda name stays stable for compatibility, but the handler now
# branches on target_format so the same worker can serve CSV, Markdown, plain
# text, and HTML table output.

import csv
import html
import io
import json as _json

import dynamo
import s3
from logger import get_logger

log = get_logger(__name__)


def _load_sheet_rows(xlsx_bytes: bytes, sheet_name: str | None):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        if sheet_name is None:
            ws = wb.worksheets[0]
        else:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"sheet not found: {sheet_name}")
            ws = wb[sheet_name]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _normalize_rows(rows):
    normalized: list[list[str]] = []
    for row in rows:
        current: list[str] = []
        for cell in row:
            current.append("" if cell is None else str(cell))
        normalized.append(current)
    return normalized


def _rows_to_csv(rows) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, dialect="excel")
    for row in _normalize_rows(rows):
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def _rows_to_text(rows) -> bytes:
    lines = ["\t".join(row).rstrip() for row in _normalize_rows(rows)]
    return "\n".join(lines).encode("utf-8")


def _rows_to_markdown(rows) -> bytes:
    normalized = _normalize_rows(rows)
    if not normalized:
        return b""
    width = max(len(row) for row in normalized)
    padded = [row + [""] * (width - len(row)) for row in normalized]
    header = padded[0]
    body = padded[1:] if len(padded) > 1 else []
    parts = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in body:
        parts.append("| " + " | ".join(row) + " |")
    return "\n".join(parts).encode("utf-8")


def _rows_to_html(rows) -> bytes:
    normalized = _normalize_rows(rows)
    parts = ["<!doctype html>", "<html>", "<body>", "<table>"]
    if normalized:
        parts.append("<thead><tr>")
        for cell in normalized[0]:
            parts.append(f"<th>{html.escape(cell)}</th>")
        parts.append("</tr></thead>")
        if len(normalized) > 1:
            parts.append("<tbody>")
            for row in normalized[1:]:
                parts.append("<tr>")
                for cell in row:
                    parts.append(f"<td>{html.escape(cell)}</td>")
                parts.append("</tr>")
            parts.append("</tbody>")
    parts.extend(["</table>", "</body>", "</html>"])
    return "\n".join(parts).encode("utf-8")


def _convert(data: bytes, body: dict) -> tuple[bytes, str]:
    params = body.get("params") or {}
    target = (params.get("target_format") or body.get("target_format") or "csv").lower()
    sheet_name = params.get("sheet") or body.get("sheet")
    rows = _load_sheet_rows(data, sheet_name)
    if target == "csv":
        return _rows_to_csv(rows), "csv"
    if target == "md":
        return _rows_to_markdown(rows), "md"
    if target == "txt":
        return _rows_to_text(rows), "txt"
    if target == "html":
        return _rows_to_html(rows), "html"
    raise ValueError("target_format must be one of: csv, html, md, txt")


def handler(event, context):
    body = _json.loads(event["Records"][0]["body"])
    job_id = body["job_id"]
    file_key = body["file_key"]

    try:
        dynamo.update_job(job_id, status="PROCESSING")
        data = s3.get_bytes(file_key)
        result, ext = _convert(data, body)
        out_key = s3.make_output_key(job_id, file_key, f"output.{ext}")
        s3.put_bytes(out_key, result)
        dynamo.mark_done(job_id, out_key)
        log.info("xlsx_to_csv done", extra={"job_id": job_id, "target_format": ext})
    except Exception as exc:
        log.exception("xlsx_to_csv failed: %s", exc)
        dynamo.mark_failed(job_id, str(exc))
        raise
